from states import logging,monthdict,Statefolder
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging

def Gujarat(data,contractor_name,contractor_address,filelocation,month,year):
    Gujaratfilespath = os.path.join(Statefolder,'Gujarat')
    logging.info('Gujarat files path is :'+str(Gujaratfilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]

    def Form_F():
        formFfilepath = os.path.join(Gujaratfilespath,'Form F Register of refusal of leave.xlsx')
        formFfile = load_workbook(filename=formFfilepath)
        logging.info('Form F file has sheet: '+str(formFfile.sheetnames))
        logging.info('create columns which are now available')

        data_formF = data.copy()
        columns=['S.no',"Unit","Unit","Address","Employee Name","Leave_due","Encash","Date_of_refusal","sign","remarks"]
        data_formF[["sign","remarks","Leave_due"]]=""
        data_formF["Date_of_refusal"]="----"
        data_formF['S.no'] = list(range(1,len(data_formF)+1))
        formF_data=data_formF[columns]
        formFsheet = formFfile['Sheet1']
        formFsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form F is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formF_data, index=False, header=False)

        logging.info('rows taken out from data')

        for r_idx, row in enumerate(rows, 7):
            for c_idx, value in enumerate(row, 1):
                formFsheet.cell(row=r_idx, column=c_idx, value=value)
                formFsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formFsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formFsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        formFfinalfile = os.path.join(filelocation,'Form F Register of refusal of leave.xlsx')
        formFfile.save(filename=formFfinalfile)




    def Form_I():
        formIfilepath = os.path.join(Gujaratfilespath,'Form I Register of employment in a shop.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')
        
        data_formI = data.copy()
        columns=["Employee Name","Gender","Age","start_time","end_time","rest_interval","mon","tue","wed","thu","Fri","sat","sun",
                                                "days_overtime","extent_of_overtime","extent_of_overtime_previously"]
    
        data_formI['S.no'] = list(range(1,len(data_formI)+1))
        
        data_formI[["mon","tue","wed","thu","Fri","sat","sun","days_overtime","extent_of_overtime","extent_of_overtime_previously"]]=""
        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']
        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')

        formIsheet.unmerge_cells("A8:P8")
        formIsheet.unmerge_cells("A9:P9")
        
        formIsheet.insert_rows(7,len(data_formI))
        
        for r_idx, row in enumerate(rows, 7):
            for c_idx, value in enumerate(row, 1):
                formIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        formIsheet.merge_cells("A"+str(8+len(data_formI))+":P"+str(8+len(data_formI)))
        formIsheet.merge_cells("A"+str(9+len(data_formI))+":P"+str(9+len(data_formI)))

        formIfinalfile = os.path.join(filelocation,'Form I Register of employment in a shop.xlsx')
        formIfile.save(filename=formIfinalfile)

    def Form_IV():
        
        formIVfilepath = os.path.join(Gujaratfilespath,'Form IV A register  of wages.xlsx')
        formIVfile = load_workbook(filename=formIVfilepath)
        logging.info('Form IV file has sheet: '+str(formIVfile.sheetnames))
        logging.info('create columns which are now available')

        data_formIV = data.copy()
        columns=['S.no',"Employee Name","Father's Name","Designation","basic","Dearness_Allowance","Earned Basic","Dearness_Allowance_2","Days Paid",
                                        "Overtime","HRA","Telephone Reimb","Bonus","Fuel Reimb","Prof Dev Reimb","Corp Attire Reimb","CCA",
                                        "FIXED MONTHLY GROSS","PF","HRA","Other Allowance","Insurance","P.Tax","Total Deductions","Net Paid",
                                        "Date of payment ","Bank A/c Number","sign"]
        
        data_formIV['S.no'] = list(range(1,len(data_formIV)+1))
        data_formIV[["Dearness_Allowance","basic","Dearness_Allowance_2","sign"]]=""
        #data_formIV["Date_overtime_worked"]=month
        formIV_data=data_formIV[columns]
        formIVsheet = formIVfile['Sheet1']
        formIVsheet.sheet_properties.pageSetUpPr.fitToPage = True
        #for column in  range(ord('A'), ord('O') + 1):
        #    formIVsheet.unmerge_cells(chr(column)+"7:"+chr(column)+"14")

        logging.info('data for form IV is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formIV_data, index=False, header=False)

        logging.info('rows taken out from data')

        for r_idx, row in enumerate(rows,11):
            for c_idx, value in enumerate(row, 1):
                formIVsheet.cell(row=r_idx, column=c_idx, value=value)
                formIVsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIVsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        formIVsheet['A3']=formIVsheet['A3'].value+" "+str(data_formIV['Unit'].unique()[0])
        formIVsheet['A4']=formIVsheet['A4'].value+" "+str(data_formIV['Unit'].unique()[0])+", "+str(data_formIV['Address'].unique()[0])
        formIVsheet['A6']="Month : "+month
        formIVfinalfile = os.path.join(filelocation,'Form IV A register  of wages.xlsx')
        formIVfile.save(filename=formIVfinalfile)


                    
    def Form_M():
        formMfilepath = os.path.join(Gujaratfilespath,'Form M Register of leave.xlsx')
        formMfile = load_workbook(filename=formMfilepath)
        logging.info('Form M file has sheet: '+str(formMfile.sheetnames))
        logging.info('create columns which are now available')

        data_formM = data.copy()
        columns=["Employee Name","Department","Date Joined","Leave Accrued","num_days","balance_days","Date Left","Date of payment "]
        data_formM[['num_days', 'balance_days']]=""
        data_formM_columns=list(data_formM.columns)
        start=data_formM_columns.index('Arrears salary')
        end=data_formM_columns.index('Total\r\nDP')
        columns.extend(data_formM_columns[start+1:end])


        formM_data=data_formM[columns]
        formMsheet = formMfile['Sheet1']

        formMsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form M is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formM_data, index=False, header=False)

        logging.info('rows taken out from data')

        def cell_write(sheet,r_idx,c_idx,value):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                sheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
            
        def start_end_date_attendance(absent_label,offset):  
            is_abs_num=0
            row_index=0
            for sheet_idx, row in enumerate(rows, 10):
                row_index=0
                for c_idx, value in enumerate(row, 1):
                    if c_idx==1:
                        try:
                            target=formMfile[value]
                        except :
                            target = formMfile.copy_worksheet(formMsheet)
                            target.title=value
                        target['A5']="Name of Employee : "+value
                        formMsheet['A4']=formMsheet['A4'].value+" "+str(data_formM['Unit'].unique()[0])+","+str(data_formM['Address'].unique()[0])    
                    elif c_idx==2:
                        target['A6']="Description of the department (if applicable):    "+value
                    elif c_idx==3:
                        target['A7']="Date of entry into service:  "+value
                    elif c_idx==4:
                        Leave_Accrued=value 
                        #cell_write(target,row_index+13,1,Leave_Accrued)
                    elif c_idx==5:
                        num_days=value
                        #cell_write(target,row_index+13,2,num_days)
                    elif c_idx==6:
                        balance_days=value
                        #cell_write(target,row_index+13,5,balance_days)
                    elif c_idx==7:
                        Date_Left=value
                        #cell_write(target,row_index+13,9,Date_Left)
                    elif c_idx==8:
                        Date_of_payment=value
                        #cell_write(target,row_index+13,10,Date_of_payment)
                    elif is_abs_num==0 and value==absent_label:
                        is_abs_num=1
                        start=columns[c_idx-1]
                        end=columns[c_idx-1]
                    elif value==absent_label:
                        is_abs_num+=1
                        end=columns[c_idx-1]
                    elif is_abs_num:
                        #target.cell(row=row_index+13, column=1+column_offset, value=is_abs_num)
                        cell_write(target,row_index+13,3+offset,start)
                        cell_write(target,row_index+13,4+offset,end)
                        cell_write(target,row_index+13,10,Date_of_payment)
                        cell_write(target,row_index+13,9,Date_Left)
                        cell_write(target,row_index+13,5,balance_days)
                        cell_write(target,row_index+13,2,num_days)
                        cell_write(target,row_index+13,1,Leave_Accrued)

                        target['F'+str(row_index+13)]="----"
                        target['G'+str(row_index+13)]="----"
                        target['H'+str(row_index+13)]="----"
                        target.insert_rows(row_index+14)
                        is_abs_num=0
                        row_index+=1
                        
                    
        absent_label="PL"
        column_offset=0           
        start_end_date_attendance(absent_label,column_offset)
        formMfile.remove(formMfile["Sheet1"])
        formMfile.remove(formMfile["Sheet2"])
        formMfile.remove(formMfile["Sheet3"])
        formMfinalfile = os.path.join(filelocation,'Form M Register of leave.xlsx')
        formMfile.save(filename=formMfinalfile)

    def Form_P():
        formPfilepath = os.path.join(Gujaratfilespath,'Form P Muster roll.xlsx')
        formPfile = load_workbook(filename=formPfilepath)
        logging.info('Form P file has sheet: '+str(formPfile.sheetnames))
        logging.info('create columns which are now available')

        data_formP = data.copy()
        
        columns=['S.no',"Employee Name","Designation","Age","Gender","Date Joined","start_time",
                                                                "end_time",'interval_for_reset_from','interval_for_reset_to']
        data_formP['interval_for_reset_to']=data_formP.rest_interval.str.split("-",expand=True)[1]
        data_formP['interval_for_reset_from']=data_formP.rest_interval.str.split("-",expand=True)[0]
        data_formP_columns=list(data_formP.columns)
        start=data_formP_columns.index('Arrears salary')
        end=data_formP_columns.index('Total\r\nDP')
        columns.extend(data_formP_columns[start+1:end])
        
        less=31-len(data_formP_columns[start+1:end])
            
        for i in range(less):
            columns.extend(["less"+str(i+1)])
            data_formP["less"+str(i+1)]=""

        data_formP['S.no'] = list(range(1,len(data_formP)+1))

        formP_data=data_formP[columns]
        formPsheet = formPfile['Sheet1']
        formPsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form P is ready')
        
        for i in range(9,20):
            formPsheet["A"+str(i)]=""

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formP_data, index=False, header=False)

        logging.info('rows taken out from data')
        
        for r_idx, row in enumerate(rows, 11):
            for c_idx, value in enumerate(row, 1):
                formPsheet.cell(row=r_idx, column=c_idx, value=value)
                formPsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formPsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formPsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        #formPsheet['AE4']=formPsheet['AE4'].value+"   "+str(data_formP['Registration_no'].unique()[0])
        formPsheet['A4']=formPsheet['A4'].value+" "+str(data_formP['Unit'].unique()[0])
        formPsheet['A5']=formPsheet['A5'].value+" "+str(data_formP['Unit'].unique()[0])
        formPsheet['A6']=formPsheet['A6'].value+" "+month

        formPfinalfile = os.path.join(filelocation,'Form P Muster roll.xlsx')
        formPfile.save(filename=formPfinalfile)

    

    def Form_Notice_holiday():
        formNotice_holidayfilepath = os.path.join(Gujaratfilespath,'Notice of holiday.xlsx')
        formNotice_holidayfile = load_workbook(filename=formNotice_holidayfilepath)
        logging.info('Form Notice_holiday file has sheet: '+str(formNotice_holidayfile.sheetnames))
        logging.info('create columns which are now available')

        data_formNotice_holiday = data.copy()
        
        columns=["Employee Name"]
        
        data_formNotice_holiday['S.no'] = list(range(1,len(data_formNotice_holiday)+1))

        formNotice_holiday_data=data_formNotice_holiday[columns]
        formNotice_holidaysheet = formNotice_holidayfile['Sheet1']
        formNotice_holidaysheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form Notice_holiday is ready')
        
        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formNotice_holiday_data, index=False, header=False)

        logging.info('rows taken out from data')
        if len(data_formNotice_holiday)>7:
            formNotice_holidaysheet.insert_rows(15,len(data_formNotice_holiday)-8)

        for r_idx, row in enumerate(rows, 14):
            for c_idx, value in enumerate(row, 1):
                formNotice_holidaysheet.cell(row=r_idx, column=c_idx, value=value)
                formNotice_holidaysheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formNotice_holidaysheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formNotice_holidaysheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                
        
        formNotice_holidaysheet['A3']=formNotice_holidaysheet['A3'].value+" "+str(data_formNotice_holiday['Unit'].unique()[0])+", "+str(data_formNotice_holiday['Address'].unique()[0])
        formNotice_holidayfinalfile = os.path.join(filelocation,'Notice of holiday.xlsx')
        formNotice_holidayfile.save(filename=formNotice_holidayfinalfile)


    Form_F()
    Form_I()
    Form_IV()
    Form_M()
    Form_P()
    Form_Notice_holiday()


        