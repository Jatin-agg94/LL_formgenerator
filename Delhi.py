from states import logging,monthdict,Statefolder
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging

def Delhi(data,contractor_name,contractor_address,filelocation,month,year):
    Delhifilespath = os.path.join(Statefolder,'Delhi')
    logging.info('Goa files path is :'+str(Delhifilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]


    #Not Complete
    def Form_G():
        formGfilepath = os.path.join(Delhifilespath,'Form G.xlsx')
        formGfile = load_workbook(filename=formGfilepath)
        logging.info('Form G file has sheet: '+str(formGfile.sheetnames))
        logging.info('create columns which are now available')

        data_formG = data.copy()

        #Part 1 form

        columns=["Employee Name","Designation","Date","start_time","end_time","interval_for_reset_from","interval_for_reset_to","Total_hrs_worked",
                                            "overtime_hrs_worked","overtime_wages_earned","Leave Type","leave_due",
                                            "leave_availed","Balance"]
        
        data_formG["Fine_damage_loss"]=data_formG["Fine"]+"\n"+data_formG["Damage or Loss"]
        data_formG[["Date",'leave_availed', 'leave_due', 'Balance','Total_hrs_worked', 'overtime_hrs_worked', 'overtime_wages_earned']]=""
        data_formG['interval_for_reset_to']=data_formG.rest_interval.str.split("-",expand=True)[1]
        data_formG['interval_for_reset_from']=data_formG.rest_interval.str.split("-",expand=True)[0]

        data_formG_columns=list(data_formG.columns)
        start=data_formG_columns.index('Arrears salary')
        end=data_formG_columns.index('Total\r\nDP')
        print(data_formG_columns,start+1,end-1)
        start_date=data_formG_columns[start+1]
        end_date=data_formG_columns[end-1]

        formG_data=data_formG[columns]
        formGsheet = formGfile['Sheet1']
        formGsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form G is ready')



        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formG_data, index=False, header=False)

        logging.info('rows taken out from data')
        added=0
        for r_idx, row in enumerate(rows, 15):
            for c_idx, value in enumerate(row, 1):
                if c_idx==1:
                    try:
                        target=formGfile[value]
                    except:
                        target = formGfile.copy_worksheet(formGsheet)
                        target.title=value
                        target["A8"]="Name of Employee "+value
                        target['A7']="Name of Establishment : "+data_formG['Unit'][0]
                        target['A4']="Year "+str(year)+"Month "+month
                        target['A5']="Wage Period:- "+start_date+"-"+end_date
                elif c_idx==2:
                    target["A9"]="Nature of Work:- "+str(value)
                else:
                    target.cell(row=15+added, column=c_idx, value=value)
                    target.cell(row=15+added, column=c_idx).font =Font(name ='Verdana', size =8)
                    target.cell(row=15+added, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=15+added, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                    #added+=1


        #Part 2 form
        columns=["Employee Name","Earned Basic","Overtime","Other Allowance","Total Earning",
                                            "Fine_damage_loss","Other Deduction","date","amount","Total Earning","Net Paid","Date of payment "
                                            ]
        data_formG[["date","amount"]]=""
        formG_data=data_formG[columns]
        formGsheet = formGfile['Sheet1']
        formGfile.remove(formGfile["Sheet1"])
        formGsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form G is ready')



        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formG_data, index=False, header=False)

        added=0
        for r_idx, row in enumerate(rows, 28):
            for c_idx, value in enumerate(row, 1):
                if c_idx==1:
                    try:
                        target=formGfile[value]
                    except:
                        target = formGfile.copy_worksheet(formGsheet)
                        print(value)
                        target.title=value
                        target["A8"]=target["A8"].value+" "+value
                        target['A7']=target['A7'].value+" : "+data_formG['Unit'][0]
                else:
                    target.cell(row=28+added, column=c_idx, value=value)
                    target.cell(row=28+added, column=c_idx).font =Font(name ='Verdana', size =8)
                    target.cell(row=28+added, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    target.cell(row=28+added, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                    #added+=1

        formGfinalfile = os.path.join(filelocation,'Form G.xlsx')
        formGfile.save(filename=formGfinalfile)



    def Form_H():
        formHfilepath = os.path.join(Delhifilespath,'Form H.xlsx')
        formHfile = load_workbook(filename=formHfilepath)
        logging.info('Form H file has sheet: '+str(formHfile.sheetnames))
        logging.info('create columns which are now available')
        def Part_I():
            data_formH = data.copy()
            columns=['S.no',"Employee Name","Designation"]

            data_formH_columns=list(data_formH.columns)
            start=data_formH_columns.index('Arrears salary')
            end=data_formH_columns.index('Total\r\nDP')
            columns.extend(data_formH_columns[start+1:end])
            
            less=31-len(data_formH_columns[start+1:end])
            for i in range(less):
                columns.extend(["less"+str(i+1)])
                data_formH["less"+str(i+1)]=""
            columns.extend(["remarks"])


            data_formH['S.no'] = list(range(1,len(data_formH)+1))
            data_formH[["remarks"]]=""
            formH_data=data_formH[columns]
            formHsheet = formHfile['Sheet1']
            formHsheet.sheet_properties.pageSetUpPr.fitToPage = True
            logging.info('data for form H is ready')

            from openpyxl.utils.dataframe import dataframe_to_rows
            rows = dataframe_to_rows(formH_data, index=False, header=False)
            rows_copy = list(dataframe_to_rows(formH_data, index=False, header=False))
            

            logging.info('rows taken out from data')
            formHsheet.unmerge_cells("A15:N15")
            formHsheet.unmerge_cells("A18:A19")
            formHsheet.unmerge_cells("B18:B19")
            formHsheet.unmerge_cells("C18:G18")
            formHsheet.unmerge_cells("H18:K18")
            formHsheet.unmerge_cells("L18:L19")
            formHsheet.unmerge_cells("M18:M19")
            formHsheet.unmerge_cells("N18:N19")
            formHsheet.insert_rows(10,len(rows_copy))
            formHsheet.delete_rows(18,2)
            formHsheet.merge_cells("C"+str(len(rows_copy)+18)+":G"+str(len(rows_copy)+18))
            formHsheet.merge_cells("H"+str(len(rows_copy)+18)+":K"+str(len(rows_copy)+18))
            formHsheet.merge_cells("A"+str(len(rows_copy)+18)+":A"+str(len(rows_copy)+19))
            formHsheet.merge_cells("B"+str(len(rows_copy)+18)+":B"+str(len(rows_copy)+19))


            for r_idx, row in enumerate(rows, 10):
                for c_idx, value in enumerate(row, 1):
                    formHsheet.cell(row=r_idx, column=c_idx, value=value)
                    formHsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                    formHsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        def Part_II():
            data_formH = data.copy()
            columns=["Employee Name","Designation",'Earned Basic','Dearness_Allowance','Other Allowance',
                                'Consolidated Salary','Overtime','Salary Advance','Fine','Other Deduction','Total Deductions',
                                'Amount_Due','sign','Date of payment ']

            
            data_formH[['Consolidated Salary']]="---"
            data_formH[["remarks",'Amount_Due','sign','Dearness_Allowance']]=""
            formH_data=data_formH[columns]
            formHsheet = formHfile['Sheet1']
            formHsheet.sheet_properties.pageSetUpPr.fitToPage = True
            logging.info('data for form H is ready')

            from openpyxl.utils.dataframe import dataframe_to_rows
            rows = dataframe_to_rows(formH_data, index=False, header=False)
            rows_copy = list(dataframe_to_rows(formH_data, index=False, header=False))
            

            logging.info('rows taken out from data')
            formHsheet.insert_rows(len(rows_copy)+20,len(rows_copy))
        
            formHsheet.merge_cells('A'+str(len(rows_copy)+16)+':A'+str(len(rows_copy)+17))
            formHsheet.merge_cells('B'+str(len(rows_copy)+16)+':B'+str(len(rows_copy)+17))
            formHsheet.merge_cells('C'+str(len(rows_copy)+16)+':G'+str(len(rows_copy)+16))
            formHsheet.merge_cells('H'+str(len(rows_copy)+16)+':K'+str(len(rows_copy)+16))
            
            for r_idx, row in enumerate(rows, len(rows_copy)+20):
                for c_idx, value in enumerate(row, 1):
                    formHsheet.cell(row=r_idx, column=c_idx, value=value)
                    formHsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                    formHsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                    border_sides = Side(style='thin')
                    formHsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

            
            data_formH_columns=list(data_formH.columns)
            
            start=data_formH_columns.index('Arrears salary')
            end=data_formH_columns.index('Total\r\nDP')

            formHsheet.merge_cells('A'+str(len(rows_copy)+13)+':N'+str(len(rows_copy)+13))
            formHsheet['A5']="Name of Establishment   "+str(data_formH['Unit'].unique()[0])
            formHsheet['H5']=formHsheet['H5'].value+"   "+str(data_formH_columns[start+1]+" "+month)
            
            formHsheet['A6']="Registration No   "+str(data_formH['Registration_no'].unique()[0])
            formHsheet['H6']=str(data_formH_columns[end-1]+" "+month)
            
            formHsheet['Q7']=str(data_formH_columns[start+1]+" "+month)
            formHsheet['U7']=str(data_formH_columns[end-1]+" "+month)
            

            formHsheet['A'+str(len(rows_copy)+14)]="Name of Establishment   "+str(data_formH['Unit'].unique()[0])
            formHsheet['A'+str(len(rows_copy)+15)]="Registration No   "+str(data_formH['Registration_no'].unique()[0])
            formHsheet['G'+str(len(rows_copy)+15)]=month
        Part_I()
        Part_II()
        
        formHfinalfile = os.path.join(filelocation,'Form H.xlsx')
        formHfile.save(filename=formHfinalfile)
        

    def Form_I_reg():
        formIfilepath = os.path.join(Delhifilespath,'Form I register of Fine.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formI = data.copy()
        columns=['S.no',"Employee Name","Father's Name","Gender","Department","nature_of_offence","cause_against_fine","FIXED MONTHLY GROSS",
                                        "Date of payment&Fine","Date of payment ","remarks"]

        data_formI['S.no'] = list(range(1,len(data_formI)+1))
        data_formI[["nature_of_offence","cause_against_fine","remarks"]]="-----"
        data_formI["Date of payment&Fine"]=data_formI["Date of payment "]+"\n"+data_formI["Fine"]
        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']
        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')

        for r_idx, row in enumerate(rows, 7):
            for c_idx, value in enumerate(row, 1):
                formIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        formIsheet['A4']=formIsheet['A4'].value+" : "+data_formI['Unit'][0]
        formIfinalfile = os.path.join(filelocation,'Form I register of Fine.xlsx')
        formIfile.save(filename=formIfinalfile)

    def Form_I():
        formIfilepath = os.path.join(Delhifilespath,'Form I.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formI = data.copy()
        columns=["Employee Name","Date Joined"]
        data_formI_columns=list(data_formI.columns)
        start=data_formI_columns.index('Arrears salary')
        end=data_formI_columns.index('Total\r\nDP')
        columns.extend(data_formI_columns[start+1:end])


        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']

        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        for column in  range(ord('A'), ord('G') + 1):
            formIsheet.unmerge_cells(chr(column)+"11:"+chr(column)+"15")
        formIsheet.unmerge_cells("H11:I15")
        formIsheet.unmerge_cells("J11:J15")
        formIsheet.unmerge_cells("K11:K15")
        logging.info('data for form I is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')

        def cell_write(sheet,r_idx,c_idx,value):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                sheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
            
        def start_end_date_attendance(absent_label,offset):  
            is_abs_num=0
            row_index=0
            for sheet_idx, row in enumerate(rows, 10):
                row_index=0
                for c_idx, value in enumerate(row, 1):
                    if c_idx==1:
                        try:
                            target=formIfile[value]
                        except :
                            target = formIfile.copy_worksheet(formIsheet)
                            target.title=value
                        target['A6']="Name of Employee : "+value
                        target['A4']="Name of Establishment : "+data_formI['Unit'][0]
                    elif c_idx==2:
                        target['A5']="Date of Employment : "+value
                    elif is_abs_num==0 and value==absent_label:
                        is_abs_num=1
                        start=columns[c_idx-1]
                        end=columns[c_idx-1]
                    elif value==absent_label:
                        is_abs_num+=1
                        end=columns[c_idx-1]
                    elif is_abs_num:
                        #target.cell(row=row_index+13, column=1+column_offset, value=is_abs_num)
                        cell_write(target,row_index+11,3+offset,start)
                        cell_write(target,row_index+11,4+offset,end)
                        cell_write(target,row_index+11,5+offset,is_abs_num)
                        is_abs_num=0
                        row_index+=1
                    
        absent_label="PL"
        column_offset=5           
        start_end_date_attendance(absent_label,column_offset)
        formIfile.remove(formIfile["Sheet1"])
        formIfinalfile = os.path.join(filelocation,'Form I.xlsx')
        formIfile.save(filename=formIfinalfile)


    def Form_II():
        formIIfilepath = os.path.join(Delhifilespath,'Form II.xlsx')
        formIIfile = load_workbook(filename=formIIfilepath)
        logging.info('Form II file has sheet: '+str(formIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formII = data.copy()
        columns=['S.no',"Employee Name","Father's Name","Gender","Department",
                                        "Damage or Loss","whether_work_showed_cause",
                                        "Date of payment & amount of deduction","num_instalments","Date of payment ","remarks"]
        
        data_formII['S.no'] = list(range(1,len(data_formII)+1))
        data_formII[["whether_work_showed_cause","num_instalments","remarks"]]="-----"
        data_formII["Date of payment & amount of deduction"]=data_formII["Date of payment "]+"\n"+data_formII["Total Deductions"]
        formII_data=data_formII[columns]
        formIIsheet = formIIfile['Sheet1']
        formIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form II is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formII_data, index=False, header=False)

        logging.info('rows taken out from data')

        for r_idx, row in enumerate(rows, 8):
            for c_idx, value in enumerate(row, 1):
                formIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        formIIsheet['A4']=formIIsheet['A4'].value+" : "+data_formII['Unit'][0]
        formIIfinalfile = os.path.join(filelocation,'Form II.xlsx')
        formIIfile.save(filename=formIIfinalfile)

    def Form_IV():
        formIVfilepath = os.path.join(Delhifilespath,'Form IV.xlsx')
        formIVfile = load_workbook(filename=formIVfilepath)
        logging.info('Form IV file has sheet: '+str(formIVfile.sheetnames))
        logging.info('create columns which are now available')

        data_formIV = data.copy()
        columns=['S.no',"Employee Name","Father's Name","Gender","Designation_Dept","Date_overtime_worked",
                                        "Extent of over-time","Total over-time","Normal hrs ",
                                        "FIXED MONTHLY GROSS","Overtime","overtime rate","ot","FIXED MONTHLY GROSS","Date of payment "]
        
        data_formIV['S.no'] = list(range(1,len(data_formIV)+1))
        data_formIV['Designation_Dept']=data_formIV["Designation"]+"_"+data_formIV["Department"]
        data_formIV[["Extent of over-time","Total over-time"]]="-----"
        data_formIV["ot"]=""
        data_formIV["Date_overtime_worked"]=month
        data_formIV["Date of payment & amount of deduction"]=data_formIV["Date of payment "]+"\n"+data_formIV["Total Deductions"]
        formIV_data=data_formIV[columns]
        formIVsheet = formIVfile['Sheet1']
        formIVsheet.sheet_properties.pageSetUpPr.fitToPage = True
        for column in  range(ord('A'), ord('O') + 1):
            formIVsheet.unmerge_cells(chr(column)+"7:"+chr(column)+"14")

        logging.info('data for form IV is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formIV_data, index=False, header=False)

        logging.info('rows taken out from data')

        for r_idx, row in enumerate(rows, 8):
            for c_idx, value in enumerate(row, 1):
                formIVsheet.cell(row=r_idx, column=c_idx, value=value)
                formIVsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIVsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        #formIVsheet['A4']=formIVsheet['A4'].value+" : "+data_formIV['Unit'][0]
        formIVsheet['A4']="Month Ending: "+month+" "+str(year)
        formIVfinalfile = os.path.join(filelocation,'Form IV.xlsx')
        formIVfile.save(filename=formIVfinalfile)
        
    Form_H()
    Form_I_reg()
    Form_I()
    Form_II()
    Form_IV()
    Form_G()

