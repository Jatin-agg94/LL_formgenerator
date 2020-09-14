# Author
# Suraj Mahangade
from states import logging,monthdict,Statefolder,count
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging
from collections import Counter

#Goa forms
def Goa(data,contractor_name,contractor_address,filelocation,month,year):
    
    Goafilespath = os.path.join(Statefolder,'Goa')
    logging.info('Goa files path is :'+str(Goafilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]
    #Form I
    def Form_I():
        #open xl file
        formIfilepath = os.path.join(Goafilespath,'Form I register of Fine.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formI = data.copy()
        #Columns accroding to mapping
        columns=['S.no',"Employee Name","Father's Name","Gender","Department","name&date_of_offence","cause_against_fine","FIXED MONTHLY GROSS",
                                        "Date of payment ","Date of payment ","remarks"]

        data_formI['S.no'] = list(range(1,len(data_formI)+1))
        #can update if when get new mapping
        data_formI[["name&date_of_offence","cause_against_fine","remarks"]]="-----"
        #take the columns from the data df
        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']
        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        #convert dataframe to rows
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')

        #loop over the rows and populate the cells
        for r_idx, row in enumerate(rows, 10):
            for c_idx, value in enumerate(row, 1):
                formIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        #unit name at column A4
        formIsheet['A4']=formIsheet['A4'].value+" : "+data_formI['Unit'][0]
        formIfinalfile = os.path.join(filelocation,'Form I register of Fine.xlsx')
        formIfile.save(filename=formIfinalfile)

    #Form II
    def Form_II():
        formIIfilepath = os.path.join(Goafilespath,'Form II register of damage or loss.xlsx')
        formIIfile = load_workbook(filename=formIIfilepath)
        logging.info('Form II file has sheet: '+str(formIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formII = data.copy()
        #Columns according to mapping
        columns=['S.no',"Employee Name","Father's Name","Gender","Department","attendancefile",
                                        "Damage or Loss","whether_work_showed_cause",
                                        "Date of payment ","num_instalments","Date of payment ","remarks"]

        data_formII['S.no'] = list(range(1,len(data_formII)+1))
        data_formII["attendancefile"]="confusing mapping"
        data_formII[["whether_work_showed_cause","num_instalments","remarks"]]="-----"
        #get the required columns
        formII_data=data_formII[columns]
        formIIsheet = formIIfile['Sheet1']
        formIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form II is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formII_data, index=False, header=False)

        logging.info('rows taken out from data')
        #loop over and populate each cell
        for r_idx, row in enumerate(rows, 10):
            for c_idx, value in enumerate(row, 1):
                formIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        #unit name
        formIIsheet['A4']=formIIsheet['A4'].value+" : "+data_formII['Unit'][0]
        formIIfinalfile = os.path.join(filelocation,'Form II register of damage or loss.xlsx')
        formIIfile.save(filename=formIIfinalfile)

    def Form_VIII():
        formVIIIfilepath = os.path.join(Goafilespath,'Form VIII register of Over time.xlsx')
        formVIIIfile = load_workbook(filename=formVIIIfilepath)
        logging.info('Form VIII file has sheet: '+str(formVIIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formVIII = data.copy()
        data_formVIII['Designation_Dept']=data_formVIII["Designation"]+"_"+data_formVIII["Department"]
        #get columns accorgin to mapping
        columns=['S.no',"Employee Name","Father's Name","Gender","Designation_Dept","attendancefile",
                                        "extent_of_overtime","total_overtime",
                                        'Normal hrs ','FIXED MONTHLY GROSS',
                                        "overtime rate",'Overtime',"ot",'CHECK CTC Gross','Date of payment ']

        data_formVIII['S.no'] = list(range(1,len(data_formVIII)+1))
        data_formVIII[["attendancefile","overtime_rate","ot"]]="Didn't find"
        data_formVIII[["extent_of_overtime","total_overtime"]]="----"
        
        formVIII_data=data_formVIII[columns]
        formVIIIsheet = formVIIIfile['Sheet1']
        formVIIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form VIII is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formVIII_data, index=False, header=False)

        logging.info('rows taken out from data')
        
        row_copy=dataframe_to_rows(formVIII_data, index=False, header=False)
        #Merge all the cells according to the form given(since only the column names are merged and the area to be populated is not) so that everything is aligned
        for i in range(len(list(row_copy))-2):
            i+=12
            formVIIIsheet.merge_cells('C'+str(i)+':D'+str(i))
            formVIIIsheet.merge_cells('F'+str(i)+':H'+str(i))
            formVIIIsheet.merge_cells('I'+str(i)+':K'+str(i))
            formVIIIsheet.merge_cells('L'+str(i)+':N'+str(i))
            formVIIIsheet.merge_cells('O'+str(i)+':R'+str(i))
            formVIIIsheet.merge_cells('S'+str(i)+':T'+str(i))
            formVIIIsheet.merge_cells('U'+str(i)+':V'+str(i))
            formVIIIsheet.merge_cells('W'+str(i)+':X'+str(i))
            formVIIIsheet.merge_cells('Y'+str(i)+':Z'+str(i))
            formVIIIsheet.merge_cells('AA'+str(i)+':AB'+str(i))
            formVIIIsheet.merge_cells('AC'+str(i)+':AD'+str(i))
            formVIIIsheet.merge_cells('AE'+str(i)+':AG'+str(i))
        
        c_idx=0
        for r_idx, row in enumerate(rows, 10):
            row_iterator=zip(row)
            while True:
                c_idx+=1
                #IF merged cell you cannot wrtie so just continue
                if type(formVIIIsheet.cell(row=r_idx, column=c_idx)).__name__ == 'MergedCell':
                    continue
                try:
                    #get the next value if not then break 
                    value=next(row_iterator)[0]
                except:
                    c_idx=0
                    break
                #populate
                formVIIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formVIIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formVIIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formVIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
                formVIIIfinalfile = os.path.join(filelocation,'Form VIII register of Over time.xlsx')
                formVIIIfile.save(filename=formVIIIfinalfile)
        
        #put month ending
        formVIIIsheet['Q4']="Month ending "+month+" "+str(year)
        formVIIIfinalfile = os.path.join(filelocation,'Form VIII register of Over time.xlsx')
        formVIIIfile.save(filename=formVIIIfinalfile)
        
    
    def From_XII():
        formXIIfilepath = os.path.join(Goafilespath,'Form XII Register of leave.xlsx')
        formXIIfile = load_workbook(filename=formXIIfilepath)
        logging.info('Form XII file has sheet: '+str(formXIIfile.sheetnames))
        #print(formXIIfile.sheetnames)
        logging.info('create columns which are now available')

        data_formXII = data.copy()
        columns=["Employee Name","Date Joined","Father's Name","Registration_no"]
        data_formXII_columns=list(data_formXII.columns)
        #Get the column locations from start of month to end of month
        start=data_formXII_columns.index('Arrears salary')
        end=data_formXII_columns.index('Total\r\nDP')
        #add the column names to list
        columns.extend(data_formXII_columns[start+1:end])


        formXII_data=data_formXII[columns]
        formXIIsheet = formXIIfile['Sheet1']

        formXIIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        #Unmerge all cells since when you add rows later it may cause problems
        formXIIsheet.unmerge_cells("A18:A19")
        formXIIsheet.unmerge_cells("B17:C17")
        formXIIsheet.unmerge_cells("D17:E17")
        formXIIsheet.unmerge_cells("B18:C18")
        formXIIsheet.unmerge_cells("D18:E18")
        formXIIsheet.unmerge_cells("F18:F19")
        formXIIsheet.unmerge_cells("G17:H17")
        formXIIsheet.unmerge_cells("G18:H18")
        formXIIsheet.unmerge_cells("I17:J17")
        formXIIsheet.unmerge_cells("I18:J18")
        
        formXIIsheet.unmerge_cells("A24:A25")
        formXIIsheet.unmerge_cells("B23:C23")
        formXIIsheet.unmerge_cells("D23:E23")
        formXIIsheet.unmerge_cells("B24:C24")
        formXIIsheet.unmerge_cells("D24:E24")
        formXIIsheet.unmerge_cells("F24:F25")
        formXIIsheet.unmerge_cells("G23:H23")
        formXIIsheet.unmerge_cells("G24:H24")
        formXIIsheet.unmerge_cells("I23:J23")
        formXIIsheet.unmerge_cells("I24:J24")

        formXIIsheet.unmerge_cells("A30:A31")
        formXIIsheet.unmerge_cells("B29:C29")
        formXIIsheet.unmerge_cells("B30:C30")
        formXIIsheet.unmerge_cells("D29:E29")
        formXIIsheet.unmerge_cells("D30:E30")
        formXIIsheet.unmerge_cells("F29:G29")
        formXIIsheet.unmerge_cells("F30:G30")

        formXIIsheet.unmerge_cells("E16:F16")
        formXIIsheet.unmerge_cells("E22:F22")
        formXIIsheet.unmerge_cells("C28:D28")
        

        logging.info('data for form I is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        #rows_copy = list(dataframe_to_rows(formXII_data, index=False, header=False))
        def cell_write(sheet,r_idx,c_idx,value):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                sheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
            
        def ELW_write(row_index,target,start,end,is_abs_num):
            cell_write(target,row_index,2,start)
            cell_write(target,row_index,3,end)
            cell_write(target,row_index,4,is_abs_num)
            #cell_write(target,row_index,5,start)
            #cell_write(target,row_index,6,end)

        def SL_write(row_index,target,start,end,is_abs_num):
            cell_write(target,row_index,2,start)
            cell_write(target,row_index,3,end)
            cell_write(target,row_index,4,is_abs_num)
            #cell_write(target,row_index,5,start)
            #cell_write(target,row_index,6,end)

        def CL_write(row_index,target,start,end,is_abs_num):
            cell_write(target,row_index,2,start)
            cell_write(target,row_index,3,end)
            cell_write(target,row_index,4,is_abs_num)
            #cell_write(target,row_index,5,start)
            #cell_write(target,row_index,6,end)
        
        def ML_write(row_index,target,start,end,is_abs_num):
            cell_write(target,row_index,2,start)
            cell_write(target,row_index,3,end)
            #cell_write(target,row_index,5,start)
            #cell_write(target,row_index,6,end)
        #there are 4 types of forms in the given form hence different function for each
        form_write={'PL':ELW_write,'SL':SL_write,'CL':CL_write,'ML':ML_write}
        
        def start_end_date_attendance(rows,absent_label,row_offset):  
           # print("infunction",row_offset)
            #num of days abs
            is_abs_num=0
            row_index=0
            #to store number of rows added for each employee
            added={}
            for sheet_idx, row in enumerate(rows, 10):
                row_index=0
                for c_idx, value in enumerate(row, 1):
                    #first column data to populate
                    if c_idx==1:
                        #check if employee sheet already exsists or not
                        try:
                            target=formXIIfile[value]
                        except:
                            target = formXIIfile.copy_worksheet(formXIIsheet)
                            #employee name as sheetname
                            target.title=value
                            #initial offset
                            row_offset[value]=14
                        #populate few cells 
                        target['A4']="Name and address of the Establishment:- "" "+str(data_formXII['Unit'].unique()[0])+", "+str(data_formXII['Address'].unique()[0])
                        target['A5']="Name of Employer:- "" "+str(data_formXII['Unit'].unique()[0])
                        target['A7']="Name of Employee : "+str(value)
                        added[target.title]=0
                    #second index data
                    elif c_idx==2:
                        target['A5']="Date of Employment : "+str(value)
                    elif c_idx==3:
                        target['A8']="Father's Name : "+str(value)
                    elif c_idx==4:
                        target['A6']="Registration No. :- "+str(value)
                    #check for absent label is present or not
                    elif is_abs_num==0 and value==absent_label:
                        is_abs_num=1
                        start=columns[c_idx-1]
                        end=columns[c_idx-1]
                    elif value==absent_label:
                        is_abs_num+=1
                        end=columns[c_idx-1]
                    elif is_abs_num:
                        #target.cell(row=row_index+13, column=1+column_offset, value=is_abs_num)
                     #   print("write",row_index,row_offset,row_index+row_offset[target.title])
                        form_write[absent_label](row_index+row_offset[target.title],target,start,end,is_abs_num)
                        target.insert_rows(row_index+row_offset[target.title]+1)
                        is_abs_num=0
                        row_index+=1
                        added[target.title]+=1

            return added
        #offset dictionary for number of rows for each employee sheet
        offset={}
        #convert to counter since it will help in adding dictionaries later
        offset=Counter(start_end_date_attendance(dataframe_to_rows(formXII_data, index=False, header=False),"PL",offset))
        
        for sheet in formXIIfile.sheetnames:
            #add offset so that we get to next part of form where we have to fill data
            offset[sheet]+=20
            #merge all the cells unmerged till the first form
            formXIIfile[sheet].merge_cells("A"+str(offset[sheet]-2)+":A"+str(offset[sheet]-1))
            formXIIfile[sheet].merge_cells("B"+str(offset[sheet]-3)+":C"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("D"+str(offset[sheet]-3)+":E"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("B"+str(offset[sheet]-2)+":C"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("D"+str(offset[sheet]-2)+":E"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("F"+str(offset[sheet]-2)+":F"+str(offset[sheet]-1))
            formXIIfile[sheet].merge_cells("G"+str(offset[sheet]-3)+":H"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("G"+str(offset[sheet]-2)+":H"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("I"+str(offset[sheet]-3)+":J"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("I"+str(offset[sheet]-2)+":J"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("A"+str(offset[sheet]-4)+":J"+str(offset[sheet]-4))
            cell_write(sheet=formXIIfile[sheet],r_idx=offset[sheet]-4,c_idx=1,value="Sick Leave")
        #add the offset with previous offset 
        offset+=Counter(start_end_date_attendance(dataframe_to_rows(formXII_data, index=False, header=False),"SL",offset))
        
        for sheet in formXIIfile.sheetnames:
            #add offset so that we get to next part of form where we have to fill data
            offset[sheet]+=6
            formXIIfile[sheet].merge_cells("A"+str(offset[sheet]-2)+":A"+str(offset[sheet]-1))
            formXIIfile[sheet].merge_cells("B"+str(offset[sheet]-3)+":C"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("D"+str(offset[sheet]-3)+":E"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("B"+str(offset[sheet]-2)+":C"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("D"+str(offset[sheet]-2)+":E"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("F"+str(offset[sheet]-2)+":F"+str(offset[sheet]-1))
            formXIIfile[sheet].merge_cells("G"+str(offset[sheet]-3)+":H"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("G"+str(offset[sheet]-2)+":H"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("I"+str(offset[sheet]-3)+":J"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("I"+str(offset[sheet]-2)+":J"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("A"+str(offset[sheet]-4)+":J"+str(offset[sheet]-4))
            cell_write(sheet=formXIIfile[sheet],r_idx=offset[sheet]-4,c_idx=1,value="Casual Leave")
        
        
        offset+=Counter(start_end_date_attendance(dataframe_to_rows(formXII_data, index=False, header=False),"CL",offset))
        
        for sheet in formXIIfile.sheetnames:
            offset[sheet]+=6
            formXIIfile[sheet].merge_cells("A"+str(offset[sheet]-2)+":A"+str(offset[sheet]-1))
            formXIIfile[sheet].merge_cells("B"+str(offset[sheet]-3)+":C"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("B"+str(offset[sheet]-2)+":C"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("D"+str(offset[sheet]-3)+":E"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("D"+str(offset[sheet]-2)+":E"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("F"+str(offset[sheet]-3)+":G"+str(offset[sheet]-3))
            formXIIfile[sheet].merge_cells("F"+str(offset[sheet]-2)+":G"+str(offset[sheet]-2))
            formXIIfile[sheet].merge_cells("A"+str(offset[sheet]-4)+":G"+str(offset[sheet]-4))
            cell_write(sheet=formXIIfile[sheet],r_idx=offset[sheet]-4,c_idx=1,value="Maternity Leave")
        offset+=Counter(start_end_date_attendance(dataframe_to_rows(formXII_data, index=False, header=False),"ML",offset))
        formXIIfinalfile = os.path.join(filelocation,'Form XII Register of leave.xlsx')
        formXIIfile.save(filename=formXIIfinalfile)
        
       
    def Form_XXI():
        formXXIfilepath = os.path.join(Goafilespath,'Form XXI Register of Employment.xlsx')
        formXXIfile = load_workbook(filename=formXXIfilepath)
        logging.info('Form XXI file has sheet: '+str(formXXIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formXXI = data.copy()
        
        columns=['S.no',"Employee Name","Father's Name","Gender","Designation","Date_of_appoinment"]
        
        data_formXXI_columns=list(data_formXXI.columns)
        start=data_formXXI_columns.index('Arrears salary')
        end=data_formXXI_columns.index('Total\r\nDP')
        #get columns of each day for present or absent
        columns.extend(data_formXXI_columns[start+1:end])
        
        less=31-len(data_formXXI_columns[start+1:end])
        for i in range(less):
            columns.extend(["less"+str(i+1)])
            data_formXXI["less"+str(i+1)]=""

        columns.extend(["normal_hours",'Overtime',"remarks"])
        data_formXXI[["normal_hours","remarks","Date_of_appoinment"]]="didn't get mapping"
        data_formXXI['S.no'] = list(range(1,len(data_formXXI)+1))

        formXXI_data=data_formXXI[columns]
        formXXIsheet = formXXIfile['Sheet1']
        formXXIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form XXI is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formXXI_data, index=False, header=False)

        logging.info('rows taken out from data')
        formXXIsheet.unmerge_cells('A23:E23')
        for r_idx, row in enumerate(rows, 14):
            for c_idx, value in enumerate(row, 1):
                formXXIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        formXXIsheet['AE4']=formXXIsheet['AE4'].value+"   "+str(data_formXXI['Registration_no'].unique()[0])
        formXXIsheet['A4']=formXXIsheet['A4'].value+" "+str(data_formXXI['Unit'].unique()[0])+", "+str(data_formXXI['Location'].unique()[0])
        formXXIsheet['A5']=formXXIsheet['A5'].value+" "+str(data_formXXI['Unit'].unique()[0])+", "+str(data_formXXI['Location'].unique()[0])
        formXXIfinalfile = os.path.join(filelocation,'Form XXI register of Over time.xlsx')
        formXXIfile.save(filename=formXXIfinalfile)



    def Form_XXIII():
        formXXIIIfilepath = os.path.join(Goafilespath,'Form XXIII Register of wages.xlsx')
        formXXIIIfile = load_workbook(filename=formXXIIIfilepath)
        logging.info('Form XXIII file has sheet: '+str(formXXIIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formXXIII = data.copy()
        #change columns
        columns=['S.no',"Employee Name","Father's Name","Designation",'Basic','Dearness_Allowance',
                                'Earned Basic','Dearness_Allowance_2','Other Allowance','Overtime',
                                 'FIXED MONTHLY GROSS','Salary Advance','PF', 'Other Deduction',
                                 'Total Deductions','Net Paid',"sign",'Date of payment ']
        
        data_formXXIII["sign"]=""
        data_formXXIII[["Basic",'Dearness_Allowance','Dearness_Allowance_2',"remarks","Date_of_appoinment"]]="didn't get mapping"
        data_formXXIII['S.no'] = list(range(1,len(data_formXXIII)+1))

        formXXIII_data=data_formXXIII[columns]
        formXXIIIsheet = formXXIIIfile['Sheet1']
        formXXIIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form XXIII is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formXXIII_data, index=False, header=False)

        logging.info('rows taken out from data')
        formXXIIIsheet.unmerge_cells('P15:R15')
        for r_idx, row in enumerate(rows, 10):
            for c_idx, value in enumerate(row, 1):
                #formXXIIIsheet.cell(row=r_idx, column=c_idx).value=value
                formXXIIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formXXIIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        
        formXXIIIsheet['P'+str(len(list(rows))+10+5)].value="Signature of Employer"
        
        formXXIIIsheet.merge_cells('P'+str(len(list(rows))+10+5)+':R'+str(len(list(rows))+10+5))
        
        formXXIIIsheet['P4']=formXXIIIsheet['P4'].value+"   "+str(data_formXXIII['Registration_no'].unique()[0])
        formXXIIIsheet['P5']=formXXIIIsheet['P5'].value+"   "+month
        formXXIIIsheet['A4']=formXXIIIsheet['A4'].value+" "+str(data_formXXIII['Unit'].unique()[0])
        formXXIIIsheet['A5']=formXXIIIsheet['A5'].value+" "+str(data_formXXIII['Unit'].unique()[0])+", "+str(data_formXXIII['Address'].unique()[0])
        
        formXXIIIfinalfile = os.path.join(filelocation,'Form XXIII Register of wages.xlsx')
        formXXIIIfile.save(filename=formXXIIIfinalfile)
        
    Form_I()
    Form_II()
    Form_VIII()
    From_XII()
    Form_XXI()
    Form_XXIII()