from states import logging,monthdict,Statefolder
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging

def Kerala():
    Keralafilespath = os.path.join(Statefolder,'Delhi')
    logging.info('Goa files path is :'+str(Keralafilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]
    def Form_A():
        formAfilepath = os.path.join(Maharashtrafilespath,'Form A muster roll.xlsx')
        formAfile = load_workbook(filename=formAfilepath)
        logging.info('Form A file has sheet: '+str(formAfile.sheetnames))
        logging.info('create columns which are now available')

        data_formA = data.copy()
        #print(sorted(data_formA.columns))
        columns=['S.no',"Emp Code","Employee Name","working_hrs_from","working_hrs_to",
                                        "interval_for_reset_from","interval_for_reset_to"]
        
        data_formA_columns=list(data_formA.columns)
        start=data_formA_columns.index('Arrears salary')
        end=data_formA_columns.index('Total\r\nDP')
        columns.extend(data_formA_columns[start+1:end])
        
        less=31-len(data_formA_columns[start+1:end])
        for i in range(less):
            columns.extend(["less"+str(i+1)])
            data_formA["less"+str(i+1)]=""

        columns.extend(["Total\r\nDP"])
        data_formA['S.no'] = list(range(1,len(data_formA)+1))
        data_formA[['interval_for_reset_to', 'working_hrs_to', 'interval_for_reset_from', 'working_hrs_from']]=""
        formA_data=data_formA[columns]
        formAsheet = formAfile['Sheet1']
        formAsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form A is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formA_data, index=False, header=False)

        logging.info('rows taken out from data')
        for r_idx, row in enumerate(rows, 10):
            for c_idx, value in enumerate(row, 1):
                formAsheet.cell(row=r_idx, column=c_idx, value=value)
                formAsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formAsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formAsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        
        formAsheet['A3']=formAsheet['A3'].value+"   "+str(data_formA['Contractor_name'].unique()[0])+","+str(data_formA['Contractor_Address'].unique()[0])
        formAsheet['A4']=formAsheet['A4'].value+" "+str(data_formA['Unit'].unique()[0])+","+str(data_formA['Address'].unique()[0])
        formAfinalfile = os.path.join(filelocation,'Form A muster roll.xlsx')
        formAfile.save(filename=formAfinalfile)