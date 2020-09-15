from states import logging,monthdict,Statefolder,count
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging
from collections import Counter


def Maharashtra(data,contractor_name,contractor_address,filelocation,month,year):
    Maharashtrafilespath = os.path.join(Statefolder,'Maharashtra')
    logging.info('Maharashtra files path is :'+str(Maharashtrafilespath))
    data.reset_index(drop=True, inplace=True)
    month_num = monthdict[month]

    def Form_I():
        formIfilepath = os.path.join(Maharashtrafilespath,'Form I register of fine.xlsx')
        formIfile = load_workbook(filename=formIfilepath)
        logging.info('Form I file has sheet: '+str(formIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formI = data.copy()
        columns=['S.no',"Employee Name","Father's Name","Gender","Department","name&date_of_offence","cause_against_fine",
                                        "FIXED MONTHLY GROSS","Date of payment ","Date of Fine","remarks"]

        data_formI['S.no'] = list(range(1,len(data_formI)+1))
        data_formI["name&date_of_offence"]="-----"
        data_formI["cause_against_fine"]="-----"
        data_formI["remarks"]=""
        data_formI["Date of Fine"]=""
        formI_data=data_formI[columns]
        formIsheet = formIfile['Sheet1']
        formIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form I is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formI_data, index=False, header=False)

        logging.info('rows taken out from data')

        for r_idx, row in enumerate(rows, 10):
            for c_idx, value in enumerate(row, 1):
                formIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        formIsheet['A5']=formIsheet['A5'].value+" : "+data_formI['Unit'][0]
        formIfinalfile = os.path.join(filelocation,'Form I register of fine.xlsx')
        formIfile.save(filename=formIfinalfile)
    
    def Form_II_Muster_Roll():
        formIIfilepath = os.path.join(Maharashtrafilespath,'Form II muster roll.xlsx')
        formIIfile = load_workbook(filename=formIIfilepath)
        logging.info('Form II file has sheet: '+str(formIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formII = data.copy()
        #print(sorted(data_formII.columns))
        columns=['S.no',"Emp Code","Employee Name","start_time","end_time",
                                        "interval_for_reset_from","interval_for_reset_to"]
        
        data_formII_columns=list(data_formII.columns)
        start=data_formII_columns.index('Arrears salary')
        end=data_formII_columns.index('Total\r\nDP')
        columns.extend(data_formII_columns[start+1:end])
        less=31-len(data_formII_columns[start+1:end])
        for i in range(less):
            columns.extend(["less"+str(i+1)])
            data_formII["less"+str(i+1)]=""

        columns.extend(["Total\r\nDP"])
        data_formII['S.no'] = list(range(1,len(data_formII)+1))
        data_formII['interval_for_reset_to']=data_formII.rest_interval.str.split("-",expand=True)[1]
        data_formII['interval_for_reset_from']=data_formII.rest_interval.str.split("-",expand=True)[0]
        formII_data=data_formII[columns]
        formIIsheet = formIIfile['Sheet1']
        formIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form II is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formII_data, index=False, header=False)

        logging.info('rows taken out from data')
        for r_idx, row in enumerate(rows, 10):
            for c_idx, value in enumerate(row, 1):
                formIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        
        formIIsheet['A3']=formIIsheet['A3'].value+"   "+str(data_formII['Contractor_name'].unique()[0])+","+str(data_formII['Contractor_Address'].unique()[0])
        formIIsheet['A4']=formIIsheet['A4'].value+" "+str(data_formII['Unit'].unique()[0])+","+str(data_formII['Address'].unique()[0])
        formIIfinalfile = os.path.join(filelocation,'Form II muster roll.xlsx')
        formIIfile.save(filename=formIIfinalfile)

    def Form_II_reg_damage_loss():
        formIIfilepath = os.path.join(Maharashtrafilespath,'Form II register of damage or losses.xlsx')
        formIIfile = load_workbook(filename=formIIfilepath)
        logging.info('Form II file has sheet: '+str(formIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formII = data.copy()
        #print(sorted(data_formII.columns))
        columns=['S.no',"Employee Name","Father's Name","Gender","Department","Damage or Loss","whether_work_showed_cause",
                                        "Date of payment & amount of deduction","num_instalments","Date of payment ","remarks"]
        
        data_formII['S.no'] = list(range(1,len(data_formII)+1))
        data_formII[["whether_work_showed_cause","num_instalments"]]="----"
        data_formII[["Date of payment & amount of deduction","remarks"]]=""
        formII_data=data_formII[columns]
        formIIsheet = formIIfile['Sheet1']
        formIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form II is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formII_data, index=False, header=False)

        logging.info('rows taken out from data')
        for r_idx, row in enumerate(rows, 9):
            for c_idx, value in enumerate(row, 1):
                formIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        
        formIIsheet['A5']="Name and Address of the Establishment "+str(data_formII['Unit'].unique()[0])+","+str(data_formII['Address'].unique()[0])
        formIIfinalfile = os.path.join(filelocation,'Form II register of damage or losses.xlsx')
        formIIfile.save(filename=formIIfinalfile)

    def Form_II_wages_reg():
        formIIfilepath = os.path.join(Maharashtrafilespath,'Form II wages register.xlsx')
        formIIfile = load_workbook(filename=formIIfilepath)
        logging.info('Form II file has sheet: '+str(formIIfile.sheetnames))
        logging.info('create columns which are now available')

        data_formII = data.copy()
        #print(sorted(data_formII.columns))
        columns=['S.no',"Emp Code","Employee Name",'Age',"Gender","Designation","Date Joined","Days Paid",
                                    "min_wages","FIXED MONTHLY GROSS","Total_Production_Piece_Rate","Overtime",
                                    "normal_wages","Earned Basic","HRA","HRA_payable","Telephone Reimb",
                                    "Bonus","Fuel Reimb","Corp Attire Reimb","CCA","Overtime","Gross_payable",
                                    "PF","P.Tax","Insurance","sal_fine_damage","Total Deductions","Net Paid",
                                    "leavefile_closeing","Monthly Increment","Leave Accrued","Closing","Date of payment ",
                                    "Bank A/c Number","Transfer_date","Net Paid","sign"]

        
        data_formII["sal_fine_damage"]=str(data_formII["Salary Advance"])+","+str(data_formII["Fine"])+","+str(data_formII["Damage or Loss"])
        
        data_formII['S.no'] = list(range(1,len(data_formII)+1))

        data_formII[["Total_Production_Piece_Rate","num_instalments"]]="----"
        data_formII[["min_wages","normal_wages","HRA_payable","Gross_payable","leavefile_closeing","Transfer_date","sign"]]=""
        formII_data=data_formII[columns]
        formIIsheet = formIIfile['Sheet1']
        formIIsheet.sheet_properties.pageSetUpPr.fitToPage = True
        logging.info('data for form II is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formII_data, index=False, header=False)

        logging.info('rows taken out from data')
        for r_idx, row in enumerate(rows, 9):
            for c_idx, value in enumerate(row, 1):
                formIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                formIIsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
        
        
        formIIsheet['A5']="Name and Address of the Establishment "+str(data_formII['Unit'].unique()[0])+","+str(data_formII['Address'].unique()[0])
        formIIfinalfile = os.path.join(filelocation,'Form II wages register.xlsx')
        formIIfile.save(filename=formIIfinalfile)

    def Form_VI_Overtime():
        formIVfilepath = os.path.join(Maharashtrafilespath,'Form IV Overtime register.xlsx')
        formIVfile = load_workbook(filename=formIVfilepath)
        logging.info('Form IV file has sheet: '+str(formIVfile.sheetnames))
        logging.info('create columns which are now available')

        data_formIV = data.copy()
        columns=['S.no',"Employee Name","Father's Name","Gender","Designation_Dept","Date_overtime_worked",
                                        "Extent of over-time","Total over-time","Normal hrs ",
                                        "FIXED MONTHLY GROSS","overtime rate","Overtime","ot","FIXED MONTHLY GROSS","Date of payment "]
        
        data_formIV['S.no'] = list(range(1,len(data_formIV)+1))
        data_formIV['Designation_Dept']=data_formIV["Designation"]+"_"+data_formIV["Department"]
        data_formIV[["Extent of over-time","Total over-time"]]="-----"
        data_formIV["ot"]=""
        data_formIV["Date_overtime_worked"]=""
        data_formIV["Date of payment & amount of deduction"]=data_formIV["Date of payment "]+"\n"+data_formIV["Total Deductions"]
        formIV_data=data_formIV[columns]
        formIVsheet = formIVfile['Sheet1']
        formIVsheet.sheet_properties.pageSetUpPr.fitToPage = True
        
        logging.info('data for form IV is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formIV_data, index=False, header=False)

        logging.info('rows taken out from data')

        for r_idx, row in enumerate(rows, 8):
            for c_idx, value in enumerate(row, 1):
                formIVsheet.cell(row=r_idx, column=c_idx, value=value)
                formIVsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIVsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        #formIVsheet['A4']=formIVsheet['A4'].value+" : "+data_formIV['Unit'][0]
        formIVsheet['A6']="Name of the Establishment : "+data_formIV['Unit'][0]
        formIVfinalfile = os.path.join(filelocation,'Form IV Overtime register.xlsx')
        formIVfile.save(filename=formIVfinalfile)


    def Form_VI_reg_advance():
        formIVfilepath = os.path.join(Maharashtrafilespath,'Form IV register of advance.xlsx')
        formIVfile = load_workbook(filename=formIVfilepath)
        logging.info('Form IV file has sheet: '+str(formIVfile.sheetnames))
        logging.info('create columns which are now available')

        data_formIV = data.copy()
        columns=['S.no',"Employee Name","Father's Name","Department","Salary Advance","purpose_advance",
                                        "num_installments_advance","Postponement_granted",
                                        "Date of payment ","remarks"]
                                        
                                        
        data_formIV['S.no'] = list(range(1,len(data_formIV)+1))
        data_formIV['Designation_Dept']=data_formIV["Designation"]+"_"+data_formIV["Department"]
        data_formIV[["purpose_advance","num_installments_advance","Postponement_granted"]]="-----"
        data_formIV["remarks"]=""
        formIV_data=data_formIV[columns]
        formIVsheet = formIVfile['Sheet1']
        formIVsheet.sheet_properties.pageSetUpPr.fitToPage = True
        
        logging.info('data for form IV is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formIV_data, index=False, header=False)

        logging.info('rows taken out from data')

        for r_idx, row in enumerate(rows, 13):
            for c_idx, value in enumerate(row, 1):
                formIVsheet.cell(row=r_idx, column=c_idx, value=value)
                formIVsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formIVsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formIVsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        #formIVsheet['A4']=formIVsheet['A4'].value+" : "+data_formIV['Unit'][0]
        formIVsheet['A6']="Name of Factory or Industrial Establishment. : "+data_formIV['Unit'][0]
        formIVfinalfile = os.path.join(filelocation,'Form IV register of advance.xlsx')
        formIVfile.save(filename=formIVfinalfile)



    def From_O():
        formOfilepath = os.path.join(Maharashtrafilespath,'Form O leave book.xlsx')
        formOfile = load_workbook(filename=formOfilepath)
        logging.info('Form O file has sheet: '+str(formOfile.sheetnames))
        #print(formOfile.sheetnames)
        logging.info('create columns which are now available')

        data_formO = data.copy()
        columns=["Employee Name","Date Joined","Department","Registration_no"]
        data_formO_columns=list(data_formO.columns)
        start_col=data_formO_columns.index('Arrears salary')
        end=data_formO_columns.index('Total\r\nDP')
        columns.extend(data_formO_columns[start_col+1:end])


        formO_data=data_formO[columns]
        formOsheet = formOfile['Sheet1']

        formOsheet.sheet_properties.pageSetUpPr.fitToPage = True

        #for column in  range(ord('A'), ord('G') + 1):
        #    formOsheet.unmerge_cells(chr(column)+"11:"+chr(column)+"15")
        formOsheet.unmerge_cells("A22:H22")
        formOsheet.unmerge_cells("A23:B23")
        formOsheet.unmerge_cells("C23:C24")
        formOsheet.unmerge_cells("D23:D24")
        formOsheet.unmerge_cells("E23:E24")
        formOsheet.unmerge_cells("F23:G24")
        formOsheet.unmerge_cells("H23:H24")
        formOsheet.unmerge_cells("F25:G25")
        formOsheet.unmerge_cells("F26:G26")
        formOsheet.unmerge_cells("F27:G27")
        
        formOsheet.unmerge_cells("A28:F28")
        formOsheet.unmerge_cells("A29:B30")
        formOsheet.unmerge_cells("C29:C31")
        formOsheet.unmerge_cells("D29:D31")
        formOsheet.unmerge_cells("E29:E31")
        formOsheet.unmerge_cells("F29:F31")
        
        
        logging.info('data for form I is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        #rows_copy = list(dataframe_to_rows(formO_data, index=False, header=False))
        def cell_write(sheet,r_idx,c_idx,value):
                sheet.cell(row=r_idx, column=c_idx, value=value)
                sheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                sheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                sheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)
            
        def PL_write(row_index,target,start,end,is_abs_num):
            cell_write(target,row_index,3,"From :"+start+" To : "+end)
            cell_write(target,row_index , 1,data_formO_columns[start_col+1])
            cell_write(target,row_index , 4,"----")
            cell_write(target,row_index , 5,"----")
            cell_write(target,row_index , 6,"----")
            cell_write(target,row_index , 7,"----")
            cell_write(target,row_index , 8,"----")
            cell_write(target,row_index , 9,str(data_formO.loc[data_formO[columns[0]]==emp_name,"Date Left"].tolist()[0]))
            #cell_write(target,row_index,4,is_abs_num)
            #cell_write(target,row_index,5,start)
            #cell_write(target,row_index,6,end)

        def FL_write(row_index,target,start,end,is_abs_num):
            cell_write(target,row_index,1,start)
            cell_write(target,row_index,2,end)
            cell_write(target,row_index, 6, "-----")
            "---------------------------------------------------------------------uncomment later---------------------------------------------------------------"
            formOfile[sheet].merge_cells("F"+str(row_index)+":G"+str(row_index))
            print("F"+str(row_index)+":G"+str(row_index))
            #cell_write(target,row_index,4,is_abs_num)
            #cell_write(target,row_index,5,start)
            #cell_write(target,row_index,6,end)
        
        def CL_write(row_index,target,start,end,is_abs_num):
            cell_write(target,row_index,2,start)
            cell_write(target,row_index,3,end)
            #cell_write(target,row_index,5,start)
            #cell_write(target,row_index,6,end)

        form_write={'PL':PL_write,'FL':FL_write,'CL':CL_write}
        
        def start_end_date_attendance(rows,absent_label,row_offset,initial_offset):  
           # print("infunction",row_offset)
            is_abs_num=0
            row_index=0
            added={}
            for sheet_idx, row in enumerate(rows, 10):
                row_index=0
                for c_idx, value in enumerate(row, 1):
                    if c_idx==1:
                        try:
                            target=formOfile[value]
                        except:
                            target = formOfile.copy_worksheet(formOsheet)
                            target.title=value
                            #initial offset
                            row_offset[value]=initial_offset
                        
                        target['A4']="Name and address of the Establishment:- "" "+str(data_formO['Unit'].unique()[0])+","+str(data_formO['Address'].unique()[0])
                        target['A5']="Name of Employer:- "" "+str(data_formO['Unit'].unique()[0])
                        target["H4"]="Name of the employer:- "+str(data_formO['Unit'].unique()[0])+"\n"+" Receipt of leave book - "
                        target['A7']="Name of Employee : "+str(value)
                        global emp_name
                        emp_name=str(value)
                        added[target.title]=0
                    elif c_idx==2:
                        target['H8']="Date of entry into service :- "+str(value)
                    elif c_idx==3:
                        target['A8']="Description of the Department (If Applicable) :-  "+str(value)
                    elif c_idx==4:
                        target['A6']="Registration No. :- "+str(value)
                    elif is_abs_num==0 and value=="PL":#absent_label:
                        is_abs_num=1
                        start=columns[c_idx-1]
                        end=columns[c_idx-1]
                    elif value=="PL":#absent_label:
                        is_abs_num+=1
                        end=columns[c_idx-1]
                    elif is_abs_num:
                        #target.cell(row=row_index+13, column=1+column_offset, value=is_abs_num)
                     #   print("write",row_index,row_offset,row_index+row_offset[target.title])
                        form_write[absent_label](row_index+row_offset[target.title],target,start,end,is_abs_num)
                        target.insert_rows(row_index+row_offset[target.title]+1)
                        is_abs_num=0
                        row_index+=1
                        added[target.title]+=1
                        
                        
                        
                        #cell_write(target,row_index , 2, str(start+1-end))
                        
                    
            return added
        offset={}
        initial_offset=13
        #for sheet in formOfile.sheetnames:
        #    offset[sheet]=initial_offset
        offset=Counter(offset)+Counter(start_end_date_attendance(dataframe_to_rows(formO_data, index=False, header=False),"PL",offset,initial_offset))
        
        for sheet in formOfile.sheetnames:
            offset[sheet]+=25
            initial_offset+=25
            formOfile[sheet].merge_cells("A"+str(offset[sheet]-3)+":H"+str(offset[sheet]-3))
            formOfile[sheet].merge_cells("A"+str(offset[sheet]-2)+":B"+str(offset[sheet]-2))
            formOfile[sheet].merge_cells("C"+str(offset[sheet]-2)+":C"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("D"+str(offset[sheet]-2)+":D"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("E"+str(offset[sheet]-2)+":E"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("F"+str(offset[sheet]-2)+":G"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("H"+str(offset[sheet]-2)+":H"+str(offset[sheet]-1))
            
        offset+=Counter(start_end_date_attendance(dataframe_to_rows(formO_data, index=False, header=False),"FL",offset,initial_offset))
        
        for sheet in formOfile.sheetnames:
            offset[sheet]+=7
            initial_offset+=7
            formOfile[sheet].merge_cells("A"+str(offset[sheet]-4)+":F"+str(offset[sheet]-4))
            formOfile[sheet].merge_cells("A"+str(offset[sheet]-3)+":B"+str(offset[sheet]-2))
            formOfile[sheet].merge_cells("C"+str(offset[sheet]-3)+":C"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("D"+str(offset[sheet]-3)+":D"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("E"+str(offset[sheet]-3)+":E"+str(offset[sheet]-1))
            formOfile[sheet].merge_cells("F"+str(offset[sheet]-3)+":F"+str(offset[sheet]-1))
            
        
        offset+=Counter(start_end_date_attendance(dataframe_to_rows(formO_data, index=False, header=False),"CL",offset,initial_offset))
        
        formOfinalfile = os.path.join(filelocation,'Form O leave book.xlsx')
        formOfile.save(filename=formOfinalfile)
    Form_I()
    Form_II_Muster_Roll()
    Form_II_reg_damage_loss()
    Form_II_wages_reg()
    Form_VI_Overtime()
    Form_VI_reg_advance()
    From_O()