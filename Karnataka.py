from states import logging,monthdict,Statefolder
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter as tk
from functools import partial
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, Side
import calendar
import logging


def Karnataka(data,contractor_name,contractor_address,filelocation,month,year):
    karnatakafilespath = os.path.join(Statefolder,'Karnataka')
    logging.info('karnataka files path is :'+str(karnatakafilespath))
    data.reset_index(drop=True, inplace=True)

    month_num = monthdict[month]

    def create_form_A():

        formAfilepath = os.path.join(karnatakafilespath,'FormA.xlsx')
        formAfile = load_workbook(filename=formAfilepath)
        logging.info('Form A file has sheet: '+str(formAfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formA = data.copy()

        data_formA['S.no'] = list(range(1,len(data_formA)+1))

        data_formA["Nationality"] =''
        data_formA['education level'] = ''
        data_formA['Category address'] = ''
        data_formA['type of employment'] = ''
        data_formA['lwf'] = ''
        data_formA['Service Book No'] = ''
        formA_columns = ["S.no",'Employee Code','Employee Name','Unit','Location',"Gender","Father's Name",'Date of Birth',"Nationality","education level",'Date Joined','Designation',"Category address","type of employment",'Mobile Tel No.','UAN Number',"PAN Number","ESIC Number","lwf","Aadhar Number","Bank A/c Number","Bank Name","Account Code","P","L","Service Book No","Date Left","Reason for Leaving"]
        formA_data = data_formA[formA_columns]


        formAsheet = formAfile['FORM A']

        formAsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form A is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formA_data, index=False, header=False)

        logging.info('rows taken out from data')

        

        for r_idx, row in enumerate(rows, 13):
            for c_idx, value in enumerate(row, 1):
                formAsheet.cell(row=r_idx, column=c_idx, value=value)
                formAsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                formAsheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                formAsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        logging.info('')

        establishment = formAsheet['L6'].value
        L6_data = establishment+' '+data_formA['Unit'][0]+', '+data_formA['Branch'][0]
        formAsheet['L6'] = L6_data

        company = formAsheet['A10'].value
        A10_data = company+' '+data_formA['Unit'][0]+', '+data_formA['Branch'][0]
        formAsheet['A10'] = A10_data

        formAfinalfile = os.path.join(filelocation,'FormA.xlsx')
        formAfile.save(filename=formAfinalfile)
        

    def create_form_B():
        formBfilepath = os.path.join(karnatakafilespath,'FormB.xlsx')
        formBfile = load_workbook(filename=formBfilepath)
        logging.info('Form B file has sheet: '+str(formBfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formB = data.copy()

        data_formB['OT hours'] = 0
        data_formB['Pay OT'] = 0
        data_formB['basic_and_allo'] = data_formB['Earned Basic']+ data_formB['Other Allowance']+ data_formB['Meal Allowance'] +data_formB['Special Allowance'] +data_formB['Personal Allowance']
        data_formB['Other EAR'] = data_formB['Other Reimb']+data_formB['Arrears']+data_formB['Other Earning']+data_formB['Variable Pay']+data_formB['Stipend'] +data_formB['Consultancy Fees']
        data_formB['VPF']=0
        data_formB['Income Tax']=0
        data_formB['EMP PF'] = data_formB['PF']
        data_formB['BankID'] = ''
        data_formB['Pay Date'] = ''
        data_formB['Remarks'] =''

        formB_columns = ['Employee Code','Employee Name','FIXED MONTHLY GROSS',	'Days Paid','OT hours',	'basic_and_allo', 'Pay OT',	'HRA',	'Tel and Int Reimb', 'Bonus', 'Fuel Reimb',	'Prof Dev Reimb', 'Corp Attire Reimb',	'CCA',	'Leave Encashment',	'Other EAR', 'Total Earning', 'PF',	'ESIC',	'VPF', 'Loan Deduction', 'Loan Interest', 'P.Tax',	'CSR',	'Income Tax', 'Insurance',	'LWF EE',	'Other Deduction',	'TDS',	'Total Deductions',	'Net Paid',	'EMP PF','BankID','Pay Date','Remarks']

        formB_data = data_formB[formB_columns]

        formBsheet = formBfile['FORM B']

        formBsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form B is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formB_data, index=False, header=False)

        logging.info('rows taken out from data')

        

        for r_idx, row in enumerate(rows, 21):
            for c_idx, value in enumerate(row, 2):
                formBsheet.cell(row=r_idx, column=c_idx, value=value)
                formBsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                border_sides = Side(style='thin')
                formBsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formBsheet['B10'].value
        B10_data = contractline+' '+contractor_name+', '+contractor_address
        formBsheet['B10'] = B10_data

        locationline = formBsheet['B11'].value
        B11_data = locationline+' '+data_formB['Unit'][0]+', '+data_formB['Branch'][0]
        formBsheet['B11'] = B11_data

        establine = formBsheet['B12'].value
        B12_data = establine+' '+data_formB['Unit'][0]+', '+data_formB['Branch'][0]
        formBsheet['B12'] = B12_data

        peline = formBsheet['B13'].value
        B13_data = peline+' '+data_formB['Unit'][0]+', '+data_formB['Branch'][0]
        formBsheet['B13'] = B13_data

        monthstart = datetime.date(year,month_num,1)
        monthend = datetime.date(year,month_num,calendar.monthrange(year,month_num)[1])
        formBsheet['B16'] = 'Wage period From: '+str(monthstart)+' to '+str(monthend)

        formBfinalfile = os.path.join(filelocation,'FormB.xlsx')
        formBfile.save(filename=formBfinalfile)

    def create_form_XXI():
        formXXIfilepath = os.path.join(karnatakafilespath,'FormXXI.xlsx')
        formXXIfile = load_workbook(filename=formXXIfilepath)
        logging.info('Form XXI file has sheet: '+str(formXXIfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formXXI = data.copy()

        data_formXXI['S.no'] = list(range(1,len(data_formXXI)+1))

        data_formXXI['a'] ='---'
        data_formXXI['b'] ='---'
        data_formXXI['c'] ='---'
        data_formXXI['e'] ='---'
        data_formXXI['f'] ='---'
        data_formXXI['g'] =''

        formXXI_columns = ['S.no','Employee Name',"Father's Name",'Designation','a','b','c','FIXED MONTHLY GROSS','e','f','g']

        formXXI_data = data_formXXI[formXXI_columns]

        formXXIsheet = formXXIfile['FORM XXI']

        formXXIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XXI is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formXXI_data, index=False, header=False)

        logging.info('rows taken out from data')


        for r_idx, row in enumerate(rows, 14):
            for c_idx, value in enumerate(row, 3):
                formXXIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                border_sides = Side(style='thin')
                formXXIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formXXIsheet['C7'].value
        C7_data = contractline+' '+contractor_name+', '+contractor_address
        formXXIsheet['C7'] = C7_data

        locationline = formXXIsheet['C8'].value
        C8_data = locationline+' '+data_formXXI['Unit'][0]+', '+data_formXXI['Branch'][0]
        formXXIsheet['C8'] = C8_data

        establine = formXXIsheet['C9'].value
        C9_data = establine+' '+data_formXXI['Unit'][0]+', '+data_formXXI['Branch'][0]
        formXXIsheet['C9'] = C9_data

        peline = formXXIsheet['C10'].value
        C10_data = peline+' '+data_formXXI['Unit'][0]+', '+data_formXXI['Branch'][0]
        formXXIsheet['C10'] = C10_data

        formXXIfinalfile = os.path.join(filelocation,'FormXXI.xlsx')
        formXXIfile.save(filename=formXXIfinalfile)



    def create_form_XXII():
        formXXIIfilepath = os.path.join(karnatakafilespath,'FormXXII.xlsx')
        formXXIIfile = load_workbook(filename=formXXIIfilepath)
        logging.info('Form XXII file has sheet: '+str(formXXIIfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formXXII = data.copy()

        data_formXXII['S.no'] = list(range(1,len(data_formXXII)+1))

        data_formXXII['b'] ='---'
        data_formXXII['c'] ='---'
        data_formXXII['d'] ='---'
        data_formXXII['e'] ='---'
        data_formXXII['f'] ='---'
        data_formXXII['g'] =''

        formXXII_columns = ['S.no','Employee Name',"Father's Name",'Designation','FIXED MONTHLY GROSS','b','c','d','e','f','g']

        formXXII_data = data_formXXII[formXXII_columns]

        formXXIIsheet = formXXIIfile['FORM XXII']

        formXXIIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XXII is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formXXII_data, index=False, header=False)

        logging.info('rows taken out from data')


        for r_idx, row in enumerate(rows, 15):
            for c_idx, value in enumerate(row, 3):
                formXXIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                border_sides = Side(style='thin')
                formXXIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formXXIIsheet['C7'].value
        C7_data = contractline+' '+contractor_name+', '+contractor_address
        formXXIIsheet['C7'] = C7_data

        locationline = formXXIIsheet['C8'].value
        C8_data = locationline+' '+data_formXXII['Unit'][0]+', '+data_formXXII['Branch'][0]
        formXXIIsheet['C8'] = C8_data

        establine = formXXIIsheet['C9'].value
        C9_data = establine+' '+data_formXXII['Unit'][0]+', '+data_formXXII['Branch'][0]
        formXXIIsheet['C9'] = C9_data

        peline = formXXIIsheet['C10'].value
        C10_data = peline+' '+data_formXXII['Unit'][0]+', '+data_formXXII['Branch'][0]
        formXXIIsheet['C10'] = C10_data

        formXXIIfinalfile = os.path.join(filelocation,'FormXXII.xlsx')
        formXXIIfile.save(filename=formXXIIfinalfile)


    def create_form_XXIII():
        formXXIIIfilepath = os.path.join(karnatakafilespath,'FormXXIII.xlsx')
        formXXIIIfile = load_workbook(filename=formXXIIIfilepath)
        logging.info('Form XXIII file has sheet: '+str(formXXIIIfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formXXIII = data.copy()

        data_formXXIII['S.no'] = list(range(1,len(data_formXXIII)+1))

        data_formXXIII['b'] ='---'
        data_formXXIII['c'] ='---'
        data_formXXIII['d'] ='---'
        data_formXXIII['e'] ='---'
        data_formXXIII['f'] ='---'
        data_formXXIII['g'] =''

        formXXIII_columns = ['S.no','Employee Name',"Father's Name",'Designation','FIXED MONTHLY GROSS','b','c','d','e','f','g']

        formXXIII_data = data_formXXIII[formXXIII_columns]

        formXXIIIsheet = formXXIIIfile['FORM XXIII']

        formXXIIIsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XXIII is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formXXIII_data, index=False, header=False)

        logging.info('rows taken out from data')


        for r_idx, row in enumerate(rows, 12):
            for c_idx, value in enumerate(row, 3):
                formXXIIIsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXIIIsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                border_sides = Side(style='thin')
                formXXIIIsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formXXIIIsheet['C5'].value
        C5_data = contractline+' '+contractor_name+', '+contractor_address
        formXXIIIsheet['C5'] = C5_data

        locationline = formXXIIIsheet['C6'].value
        C6_data = locationline+' '+data_formXXIII['Unit'][0]+', '+data_formXXIII['Branch'][0]
        formXXIIIsheet['C6'] = C6_data

        establine = formXXIIIsheet['C7'].value
        C7_data = establine+' '+data_formXXIII['Unit'][0]+', '+data_formXXIII['Branch'][0]
        formXXIIIsheet['C7'] = C7_data

        peline = formXXIIIsheet['C8'].value
        C8_data = peline+' '+data_formXXIII['Unit'][0]+', '+data_formXXIII['Branch'][0]
        formXXIIIsheet['C8'] = C8_data

        formXXIIIfinalfile = os.path.join(filelocation,'FormXXIII.xlsx')
        formXXIIIfile.save(filename=formXXIIIfinalfile)


    def create_form_XX():
        formXXfilepath = os.path.join(karnatakafilespath,'FormXX.xlsx')
        formXXfile = load_workbook(filename=formXXfilepath)
        logging.info('Form XX file has sheet: '+str(formXXfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_formXX = data.copy()

        data_formXX['S.no'] = list(range(1,len(data_formXX)+1))

        data_formXX['a'] ='---'
        data_formXX['b'] ='---'
        data_formXX['c'] ='---'
        data_formXX['d'] ='---'
        data_formXX['e'] ='---'
        data_formXX['f'] ='---'
        data_formXX['g'] ='---'
        data_formXX['h'] ='---'
        data_formXX['i'] =''

        formXX_columns = ['S.no','Employee Name',"Father's Name",'Designation','a','b','c','d','e','f','g','h','i']

        formXX_data = data_formXX[formXX_columns]

        formXXsheet = formXXfile['FORM XX']

        formXXsheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for form XX is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(formXX_data, index=False, header=False)

        logging.info('rows taken out from data')


        for r_idx, row in enumerate(rows, 15):
            for c_idx, value in enumerate(row, 3):
                formXXsheet.cell(row=r_idx, column=c_idx, value=value)
                formXXsheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                border_sides = Side(style='thin')
                formXXsheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = formXXsheet['C6'].value
        C6_data = contractline+' '+contractor_name+', '+contractor_address
        formXXsheet['C6'] = C6_data

        locationline = formXXsheet['C7'].value
        C7_data = locationline+' '+data_formXX['Unit'][0]+', '+data_formXX['Branch'][0]
        formXXsheet['C7'] = C7_data

        establine = formXXsheet['C8'].value
        C8_data = establine+' '+data_formXX['Unit'][0]+', '+data_formXX['Branch'][0]
        formXXsheet['C8'] = C8_data

        peline = formXXsheet['C9'].value
        C9_data = peline+' '+data_formXX['Unit'][0]+', '+data_formXX['Branch'][0]
        formXXsheet['C9'] = C9_data

        formXXfinalfile = os.path.join(filelocation,'FormXX.xlsx')
        formXXfile.save(filename=formXXfinalfile)

    def create_wages():
        wagesfilepath = os.path.join(karnatakafilespath,'Wages.xlsx')
        wagesfile = load_workbook(filename=wagesfilepath)
        logging.info('wages file has sheet: '+str(wagesfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_wages = data.copy()

        data_wages['S.no'] = list(range(1,len(data_wages)+1))

        data_wages['Site Address'] = ''

        data_wages['OT hours'] = 0
        data_wages['Conv'] = 0
        data_wages['Sub Allow'] = 0
        data_wages['Fines']=0
        data_wages['Damages']=0
        data_wages['Pay mode'] = ''
        data_wages['Remarks'] =''

        wages_columns = ['S.no','Employee Code','Employee Name','Gender','Designation','Department','Site Address','Date Joined','ESIC Number','PF Number','FIXED MONTHLY GROSS','Days Paid','OT hours','Earned Basic','HRA','Conv','Medical Allowance','Tel and Int Reimb', 'Bonus', 'Fuel Reimb',	'Prof Dev Reimb', 'Corp Attire Reimb','Special Allowance',	'CCA','Other Earning',	'Sub Allow','Leave Encashment', 'Total Earning','ESIC', 'PF','P.Tax','TDS','CSR','Insurance','Salary Advance','Fines','Damages','Other Deduction',	'Total Deductions',	'Net Paid','Pay mode','Bank A/c Number','Remarks']

        wages_data = data_wages[wages_columns]

        wagessheet = wagesfile['Wages']

        wagessheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for wages is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(wages_data, index=False, header=False)

        logging.info('rows taken out from data')

        

        for r_idx, row in enumerate(rows, 21):
            for c_idx, value in enumerate(row, 2):
                wagessheet.cell(row=r_idx, column=c_idx, value=value)
                wagessheet.cell(row=r_idx, column=c_idx).font =Font(name ='Verdana', size =8)
                border_sides = Side(style='thin')
                wagessheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        contractline = wagessheet['A10'].value
        A10_data = contractline+' '+contractor_name+', '+contractor_address
        wagessheet['A10'] = A10_data

        locationline = wagessheet['A11'].value
        A11_data = locationline+' '+data_wages['Unit'][0]+', '+data_wages['Branch'][0]
        wagessheet['A11'] = A11_data

        establine = wagessheet['A12'].value
        A12_data = establine+' '+data_wages['Unit'][0]+', '+data_wages['Branch'][0]
        wagessheet['A12'] = A12_data

        peline = wagessheet['A13'].value
        A13_data = peline+' '+data_wages['Unit'][0]+', '+data_wages['Branch'][0]
        wagessheet['A13'] = A13_data

        wagessheet['F4'] = 'Combined Muster Roll-cum-Register of Wages in lieu of '+month+' '+str(year)

        wagesfinalfile = os.path.join(filelocation,'Wages.xlsx')
        wagesfile.save(filename=wagesfinalfile)

    
    def create_form_H_F(form):
        if form=='FORM H':
            formHfilepath = os.path.join(karnatakafilespath,'FormH.xlsx')
        if form=='FORM F':
            formHfilepath = os.path.join(karnatakafilespath,'FormF.xlsx')
        formHfile = load_workbook(filename=formHfilepath)
        logging.info('file has sheet: '+str(formHfile.sheetnames))
        sheetformh = formHfile[form]

        
        logging.info('create columns which are now available')

        data_formH = data.copy()

        def attandance_data(employee_attendance,i):

            leavelist = list(employee_attendance.columns[(employee_attendance=='PL').iloc[i]])
            empcodeis = employee_attendance.iloc[i]['Employee Code']
            logging.info(empcodeis)
            if 'Leave Type' in leavelist:
                leavelist.remove('Leave Type')
            emp1 = pd.DataFrame(leavelist)
            
            
            if len(emp1.index)==0:
                defaultemp = {'emp':(employee_attendance).iloc[i]['Employee Code'],'startdate':0,'enddate':0,'days':0,'start_date':'-------','end_date':'-------'}
                emp1 = pd.DataFrame(defaultemp, index=[0])
                emp1.index = np.arange(1, len(emp1) + 1)
                emp1['s.no'] = emp1.index
                emp1.reset_index(drop=True, inplace=True)
                emp1['from'] = datetime.date(year,month_num,1)
                emp1['to'] = datetime.date(year,month_num,calendar.monthrange(year,month_num)[1])
                emp1['totaldays'] = calendar.monthrange(year,month_num)[1]
                emp1['leavesearned'] = (employee_attendance).iloc[i]['Monthly Increment']
                emp1['leavesstart'] = float(employee_attendance.iloc[i]['Opening'])
                emp1['leavesend'] = emp1.leavesstart - emp1.days
                emp1 = emp1[["s.no","from","to","totaldays","leavesearned","leavesstart","start_date","end_date","days","leavesend"]]
            else:
                logging.info(emp1)
                emp1.columns = ['Leaves']
                emp1['emp'] = (employee_attendance).iloc[i]['Employee Code']
                emp1['Leavesdays'] = emp1.Leaves.str[5:7].astype(int)
                emp1['daysdiff'] = (emp1.Leavesdays.shift(-1) - emp1.Leavesdays).fillna(0).astype(int)
                emp1['startdate'] = np.where(emp1.daysdiff.shift() != 1, emp1.Leavesdays, 0)
                emp1['enddate'] = np.where(emp1.daysdiff!=1, emp1.Leavesdays, 0)
                emp1.drop(emp1[(emp1.startdate==0) & (emp1.enddate==0)].index, inplace=True)
                emp1['startdate'] = np.where(emp1.startdate ==0, emp1.startdate.shift(), emp1.startdate).astype(int)
                emp1['enddate'] = np.where(emp1.enddate ==0, emp1.enddate.shift(-1), emp1.enddate).astype(int)
                emp1 = emp1[['emp','startdate','enddate']]
                emp1.drop_duplicates(subset='startdate', inplace=True)
                emp1['days'] = emp1.enddate -emp1.startdate +1
                emp1['start_date'] = [datetime.date(year,month_num,x) for x in emp1.startdate]
                emp1['end_date'] = [datetime.date(year,month_num,x) for x in emp1.enddate]
                emp1.index = np.arange(1, len(emp1) + 1)
                emp1['s.no'] = emp1.index
                emp1.reset_index(drop=True, inplace=True)
                emp1['from'] = datetime.date(year,month_num,1)
                emp1['to'] = datetime.date(year,month_num,calendar.monthrange(year,month_num)[1])
                emp1['totaldays'] = calendar.monthrange(year,month_num)[1]
                emp1['leavesearned'] = (employee_attendance).iloc[i]['Monthly Increment']
                emp1['totalleaves']= float(employee_attendance.iloc[i]['Opening'])
                emp1['cumdays']=emp1['days'].cumsum()
                emp1['leavesend'] = emp1.totalleaves - emp1.cumdays
                emp1['leavesstart'] =emp1['totalleaves']
                emp1 = emp1[["s.no","from","to","totaldays","leavesearned","leavesstart","start_date","end_date","days","leavesend"]]
                
            
            return emp1

        def prepare_emp_sheet(emp1,sheet_key,key,name,fathername):
            
            sheet1 = formHfile.copy_worksheet(sheetformh)
            sheet1.title = sheet_key
            lastline = sheet1['B18'].value
            sheet1['B18'] =''

            if len(emp1)>3:
                lastlinerow = 'B'+str(18+len(emp1))
            else:
                lastlinerow = 'B18'

            
            logging.info(lastlinerow)
            sheet1[lastlinerow] = lastline

            
            from openpyxl.utils.dataframe import dataframe_to_rows
            rows = dataframe_to_rows(emp1, index=False, header=False)

            for r_idx, row in enumerate(rows, 14):
                for c_idx, value in enumerate(row, 2):
                    sheet1.cell(row=r_idx, column=c_idx, value=value)
            sheet1['H5']=key
            sheet1['F7']=name
            sheet1['F8']=fathername

            sheet1.sheet_properties.pageSetUpPr.fitToPage = True

        emp_count = len(data_formH.index)
        emp_dic = dict()
        for i in range(0,emp_count):
            key = (data_formH).iloc[i]['Employee Code']
            emp_dic[key] = attandance_data(data_formH,i)
            sheet_key = form+'_'+str(key)
            name= data_formH[data_formH['Employee Code']==key]['Employee Name'].values[0]
            fathername= data_formH[data_formH['Employee Code']==key]["Father's Name"].values[0]
            logging.info(name)
            logging.info(fathername)
            prepare_emp_sheet(emp_dic[key],sheet_key,key,name,fathername)
            logging.info(key)
            logging.info(sheet_key)
        if form=='FORM H':
            formHfinalfile = os.path.join(filelocation,'FormH.xlsx')
        if form=='FORM F':
            formHfinalfile = os.path.join(filelocation,'FormF.xlsx')
        
        formHfile.remove(sheetformh)
        formHfile.save(filename=formHfinalfile)

    
    def create_muster():

        musterfilepath = os.path.join(karnatakafilespath,'Muster.xlsx')
        musterfile = load_workbook(filename=musterfilepath)
        logging.info('muster file has sheet: '+str(musterfile.sheetnames))

        
        logging.info('create columns which are now available')

        data_muster = data.copy()

        data_muster['S.no'] = list(range(1,len(data_muster)+1))

        first3columns = ["S.no",'Employee Code','Employee Name']
        last2columns = ["Date Left","Days Paid"]

        columnstotake =[]
        days = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','29','30','31']
        for day in days:
            for col in data_muster.columns:
                if col[5:7]==day:
                    columnstotake.append(col)
        if len(columnstotake)==28:

            columnstotake.append('29')
            columnstotake.append('30')
            columnstotake.append('31')
            data_muster['29'] = ''
            data_muster['30'] = ''
            data_muster['31'] = ''
            
        elif len(columnstotake)==29:
            columnstotake.append('30')
            columnstotake.append('31')
            data_muster['30'] = ''
            data_muster['31'] = ''

        elif len(columnstotake)==30:
            columnstotake.append('31')
            data_muster['31'] = ''

        muster_columns = first3columns+columnstotake+last2columns

        muster_data = data_muster[muster_columns]

        mustersheet = musterfile['Muster']

        mustersheet.sheet_properties.pageSetUpPr.fitToPage = True

        logging.info('data for muster is ready')

        from openpyxl.utils.dataframe import dataframe_to_rows
        rows = dataframe_to_rows(muster_data, index=False, header=False)

        logging.info('rows taken out from data')

        

        for r_idx, row in enumerate(rows, 18):
            for c_idx, value in enumerate(row, 2):
                mustersheet.cell(row=r_idx, column=c_idx, value=value)
                mustersheet.cell(row=r_idx, column=c_idx).font =Font(name ='Bell MT', size =10)
                mustersheet.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                border_sides = Side(style='thin')
                mustersheet.cell(row=r_idx, column=c_idx).border = Border(outline= True, right=border_sides, bottom=border_sides)

        logging.info('')

        contractline = mustersheet['B10'].value
        B10_data = contractline+' '+contractor_name+', '+contractor_address
        mustersheet['B10'] = B10_data

        locationline = mustersheet['B11'].value
        B11_data = locationline+' '+data_muster['Unit'][0]+', '+data_muster['Branch'][0]
        mustersheet['B11'] = B11_data

        establine = mustersheet['B12'].value
        B12_data = establine+' '+data_muster['Unit'][0]+', '+data_muster['Branch'][0]
        mustersheet['B12'] = B12_data

        peline = mustersheet['B13'].value
        B13_data = peline+' '+data_muster['Unit'][0]+', '+data_muster['Branch'][0]
        mustersheet['B13'] = B13_data

        mustersheet['B4'] = 'Combined Muster Roll-cum-Register of Wages in lieu of '+month+' '+str(year)

        musterfinalfile = os.path.join(filelocation,'Muster.xlsx')
        musterfile.save(filename=musterfinalfile)

    def create_formXIX():

        formXIXfilepath = os.path.join(karnatakafilespath,'FormXIX.xlsx')
        formXIXfile = load_workbook(filename=formXIXfilepath)
        logging.info('Form XIX file has sheet: '+str(formXIXfile.sheetnames))
        sheetformXIX = formXIXfile['FORM XIX']

        
        logging.info('create columns which are now available')

        data_formXIX = data.copy()

        emp_count = len(data_formXIX.index)
        
        for i in range(0,emp_count):
            key = (data_formXIX).iloc[i]['Employee Code']
            sheet_key = 'FORM XIX_'+str(key)

            emp_data = (data_formXIX).iloc[i]
            emp_data.fillna(value='', inplace=True)

            sheet1 = formXIXfile.copy_worksheet(sheetformXIX)
            sheet1.title = sheet_key
            sheet1['D7'] = contractor_name+', '+contractor_address
            sheet1['D8'] = emp_data['Unit']+', '+emp_data['Branch']
            sheet1['D9'] = emp_data['Unit']+', '+emp_data['Branch']
            sheet1['D10'] = emp_data['Unit']+', '+emp_data['Branch']
            sheet1['D11'] = emp_data['Employee Name']
            sheet1['D12'] = emp_data['Gender']
            sheet1['D13'] = month+'-'+str(year)
            sheet1['D14'] = key
            sheet1['D15'] = emp_data['Days Paid']
            sheet1['D16'] = emp_data['Earned Basic']
            sheet1['D17'] = emp_data['HRA']
            sheet1['D18'] = emp_data['Tel and Int Reimb']
            sheet1['D19'] = emp_data['Bonus']
            sheet1['D20'] = emp_data['Fuel Reimb']
            sheet1['D21'] = emp_data['Corp Attire Reimb']
            sheet1['D22'] = emp_data['CCA']
            sheet1['D23'] = emp_data['Other Earning']
            sheet1['D24'] = emp_data['Total Earning']
            sheet1['D25'] = emp_data['Insurance']
            sheet1['D26'] = emp_data['P.Tax']
            sheet1['D27'] = emp_data['TDS']
            sheet1['D28'] = emp_data['Total Deductions']
            sheet1['D29'] = emp_data['Net Paid']

        formXIXfinalfile = os.path.join(filelocation,'FormXIX.xlsx')
        formXIXfile.remove(sheetformXIX)
        formXIXfile.save(filename=formXIXfinalfile)

    def create_ecard():

        ecardfilepath = os.path.join(karnatakafilespath,'Employment card.xlsx')
        ecardfile = load_workbook(filename=ecardfilepath)
        logging.info('Employment card file has sheet: '+str(ecardfile.sheetnames))
        sheetecard = ecardfile['Employment card']

        
        logging.info('create columns which are now available')

        data_ecard = data.copy()

        emp_count = len(data_ecard.index)
        
        for i in range(0,emp_count):
            key = (data_ecard).iloc[i]['Employee Code']
            sheet_key = 'Employment card_'+str(key)

            emp_data = (data_ecard).iloc[i]
            emp_data.fillna(value='', inplace=True)

            sheet1 = ecardfile.copy_worksheet(sheetecard)
            sheet1.title = sheet_key
            sheet1['B4'] = contractor_name
            sheet1['B5'] = ''
            sheet1['B6'] = ''
            sheet1['B7'] = ''
            sheet1['B8'] = emp_data['Department']
            sheet1['B9'] = contractor_address
            sheet1['B10'] = emp_data['Unit']
            sheet1['B11'] = emp_data['Registration_no']
            sheet1['B12'] = ''
            sheet1['B13'] = ''
            sheet1['B14'] = emp_data['Employee Name']
            sheet1['B15'] = emp_data['Aadhar Number']
            sheet1['B16'] = emp_data['Mobile Tel No.']
            sheet1['B17'] = key
            sheet1['B18'] = emp_data['Designation']
            sheet1['B19'] = emp_data['Net Paid']
            sheet1['B20'] = emp_data['Date Joined']
            

        ecardfinalfile = os.path.join(filelocation,'Employment card.xlsx')
        ecardfile.remove(sheetecard)
        ecardfile.save(filename=ecardfinalfile)
            




        
    
    create_form_A()
    create_form_B()
    create_form_XXI()
    create_form_XXII()
    create_form_XXIII()
    create_form_XX()
    create_wages()
    create_form_H_F('FORM H')
    create_form_H_F('FORM F')
    create_muster()
    create_formXIX()
    create_ecard()